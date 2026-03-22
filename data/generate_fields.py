import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "OFF Fields"

CAT_COLORS = {
    "Identity & Display":      "D6EAF8",
    "Ingredients & Additives": "D5F5E3",
    "Scoring":                 "FEF9E7",
    "Nutrition":               "FDEDEC",
    "Allergens":               "F4ECF7",
    "Certifications & Labels": "E8F8F5",
    "Images":                  "FDF2E9",
    "Geographic & Retail":     "EBF5FB",
    "Data Quality":            "F2F3F4",
    "Metadata":                "FDFEFE",
}

# (category, api_field_path, description)
fields = [
    ("Identity & Display", "code",
     'The barcode number on the product. e.g. "8901719134845"'),
    ("Identity & Display", "product_name",
     'The name of the product. e.g. "parle"'),
    ("Identity & Display", "brands",
     'The brand that makes the product. e.g. "Parle"'),
    ("Identity & Display", "quantity",
     'Raw pack size text as printed on the label. e.g. "45gm"'),
    ("Identity & Display", "product_quantity",
     'Parsed numeric quantity extracted from the quantity field. e.g. 45'),
    ("Identity & Display", "product_quantity_unit",
     'Normalised unit paired with product_quantity. e.g. "g"'),
    ("Identity & Display", "categories_tags",
     'Product category tags. e.g. ["en:snacks", "en:biscuits"]'),
    ("Identity & Display", "url",
     'Link to the product page on Open Food Facts. e.g. "https://world.openfoodfacts.org/product/8901719134845/parle"'),

    ("Ingredients & Additives", "ingredients_text",
     'Full ingredient list as printed on the pack. e.g. "REFINED WHEAT FLOUR (MAIDA) 68%, SUGAR, REFINED PALM OIL..."'),
    ("Ingredients & Additives", "additives_tags",
     'List of additives as E-number codes. e.g. ["en:e330", "en:e472e", "en:e500"]'),
    ("Ingredients & Additives", "additives_n",
     'Total count of additives found. e.g. 4'),

    ("Scoring", "nutrition_grade_fr",
     'Nutri-Score grade from A (best) to E (worst). e.g. "e"'),
    ("Scoring", "nova_group",
     'Food processing level from 1 (unprocessed) to 4 (ultra-processed). e.g. 4'),
    ("Scoring", "ecoscore_grade",
     'Environmental impact grade from A (best) to E (worst). e.g. "d"'),

    ("Nutrition", "serving_size",
     'Declared serving size as printed on pack. e.g. "30g" (null for Parle G)'),
    ("Nutrition", "nutriments.energy-kcal_100g",
     'Calories per 100g. e.g. 454'),
    ("Nutrition", "nutriments.proteins_100g",
     'Protein in grams per 100g. e.g. 6.9'),
    ("Nutrition", "nutriments.carbohydrates_100g",
     'Total carbohydrates in grams per 100g. e.g. 77.3'),
    ("Nutrition", "nutriments.sugars_100g",
     'Sugar in grams per 100g. e.g. 25.5'),
    ("Nutrition", "nutriments.fat_100g",
     'Total fat in grams per 100g. e.g. 13'),
    ("Nutrition", "nutriments.saturated-fat_100g",
     'Saturated fat in grams per 100g. e.g. 6'),
    ("Nutrition", "nutriments.fiber_100g",
     'Dietary fibre in grams per 100g. e.g. null (not filled for Parle G)'),
    ("Nutrition", "nutriments.sodium_100g",
     'Sodium in grams per 100g. e.g. 0.296'),
    ("Nutrition", "nutriments.salt_100g",
     'Salt in grams per 100g. e.g. 0.74'),

    ("Allergens", "allergens_tags",
     'Confirmed allergens in the product. e.g. ["en:gluten", "en:milk"]'),
    ("Allergens", "traces_tags",
     '"May contain" allergen warnings. e.g. [] (none for Parle G)'),

    ("Certifications & Labels", "labels_tags",
     'Certifications printed on the pack (organic, vegan, halal, jain, fssai, etc.). e.g. [] (none for Parle G)'),
    ("Certifications & Labels", "packaging_materials_tags",
     'Packaging material type (plastic, glass, cardboard, etc.). e.g. [] (none for Parle G)'),

    ("Images", "image_front_url",
     'URL of the front-of-pack photo. e.g. "https://images.openfoodfacts.org/images/products/890/.../front_en.11.400.jpg"'),
    ("Images", "selected_images.ingredients.display.en",
     'URL of the ingredients label photo. Only present if a photo was uploaded. e.g. null for Parle G'),
    ("Images", "selected_images.nutrition.display.en",
     'URL of the nutrition table photo. Only present if a photo was uploaded. e.g. null for Parle G'),

    ("Geographic & Retail", "stores",
     'Stores where the product is sold. e.g. "" (blank for Parle G)'),
    ("Geographic & Retail", "manufacturing_places",
     'Where the product is made. e.g. "India"'),
    ("Geographic & Retail", "languages_tags",
     'Languages the product info is available in. e.g. ["en:english"]'),

    ("Data Quality", "completeness",
     'How complete the product entry is, from 0 to 1. e.g. 0.7'),
    ("Data Quality", "data_quality_errors_tags",
     'Validation errors flagged by OFF (e.g. nutrients don\'t add up). e.g. []'),
    ("Data Quality", "states_tags",
     'Which parts of the product are complete or still missing. e.g. ["en:nutrition-facts-completed", "en:ingredients-completed"]'),

    ("Metadata", "last_modified_t",
     'Unix timestamp of the last edit on OFF. e.g. 1768975293'),
    ("Metadata", "created_at  (our own)",
     'Timestamp we set when we import the record. Not from OFF.'),
]

def hex_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

headers    = ["Category", "OFF API Field Path", "Description", "Include?"]
col_widths = [22, 36, 76, 12]

for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.font      = Font(bold=True, color="FFFFFF", size=11)
    cell.fill      = hex_fill("1A1A2E")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = border()
    ws.column_dimensions[get_column_letter(col)].width = w

ws.row_dimensions[1].height = 28

prev_cat = None
for i, (cat, field, desc) in enumerate(fields, start=2):
    bg = CAT_COLORS.get(cat, "FFFFFF")

    cat_cell = ws.cell(row=i, column=1, value=cat if cat != prev_cat else "")
    cat_cell.fill      = hex_fill(bg)
    cat_cell.font      = Font(bold=(cat != prev_cat), size=10)
    cat_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cat_cell.border    = border()
    prev_cat = cat

    for col, val in enumerate([field, desc], start=2):
        cell = ws.cell(row=i, column=col, value=val)
        cell.fill      = hex_fill(bg)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border    = border()

    chk = ws.cell(row=i, column=4, value=True)
    chk.alignment = Alignment(horizontal="center", vertical="center")
    chk.border    = border()

    ws.row_dimensions[i].height = 50

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:D{len(fields)+1}"

out = "/Users/sauravkrishna/Ingredio/data/OFF_fields.xlsx"
wb.save(out)
print(f"Saved → {out}")
